import twitchio
import openai
import openpyxl
import time

class AnimeOpeningBlindTest:
    def __init__(self, bot_token, client_id, client_secret, bot_username, channel_name, excel_file_path, openai_api_key):
        # Initialize Twitch bot
        self.bot = twitchio.Bot(
            token=bot_token,
            client_id=client_id,
            client_secret=client_secret,
            nick=bot_username,
            prefix='!',
            initial_channels=[channel_name]
        )
        
        # Load Excel spreadsheet and get list of anime openings
        self.wb = openpyxl.load_workbook(excel_file_path)
        self.openings_ws = self.wb.active
        self.openings = []
        for row in self.openings_ws.iter_rows(min_row=2, min_col=1, values_only=True):
            self.openings.append(row[0])
        
        # Initialize OpenAI API
        openai.api_key = openai_api_key
        
        # Initialize score sheet
        self.scores_ws = self.wb.create_sheet("Scores")
        self.scores_ws.append(["Username", "Score"])
        self.wb.save(excel_file_path)
        
        # Set up command for viewers to join the blind test
        @self.bot.command(name='join')
        async def join(ctx):
            # Get the viewer's username
            username = ctx.author.name
            # Add the username to the score sheet
            self.scores_ws.append([username, 0])
            self.wb.save(excel_file_path)
            # Send a message confirming the viewer has joined
            await ctx.send(f"{username}, you have joined the blind test!")
            
        # Start the bot
        self.bot.run()
    
    async def start_round(self):
        # Choose a random anime opening
        opening = self.openings.pop(0)
        # Play the opening
        await self.bot.send(f"Guess the name of this anime opening: {opening}")
        # Wait for 20 seconds
        time.sleep(20)
        # Check answers and update scores
        for msg in self.bot._ws.messages:
            if msg.type == "TWITCHCHATMESSAGE" and msg.content.startswith("!"):
                continue
            answer = msg.content
            username = msg.author.name
            if self.is_correct_answer(answer, opening):
                # Award 5 points for a correct answer
                score = self.get_score(username) + 5
                self.set_score(username, score)
                await self.bot.send(f"Congratulations {username}, you got the correct answer!")
        # Go to the next round
        if len(self.openings) > 0:
            await self.start_round()
        else:
            # End of game
            await self.bot.send("The blind test is over! Here are the final scores:")
            for row in self.scores_ws.iter_rows(min_row=2, min_col=1, values_only=True):
                username = row[0]
                score = row[1]
                await self.bot.send(f"{username}: {score} points")
    
    def is_correct_answer(self, answer, correct_answer):
        # Check if answer matches the correct answer, with some leeway for typos and differences in formatting
        response = openai.Completion.create(
            engine="text-davinci-002",
            prompt=f"Is \"{answer}\" the correct answer to the anime opening \"{correct_answer}\"?",
            temperature=0.5,
            max_tokens=1,
        )
        return response.choices[0].text.lower().strip() == "yes"